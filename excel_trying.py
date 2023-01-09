from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
wb = Workbook()
wb=load_workbook(filename='price.xlsx')
#sheet = wb['TDSheet']
sheet = wb.active
good=[]
for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=1):
        #good=[]
        for cell in row:
            good.append(cell.value)
            #print(cell.value,end=' ')
        #print()
user='КАНЦТОВАРЫ'


#print(good)
#print(sheet.cell(row=1,column=1).value)
r2=set()
j=0
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
    for cell in row:
        j+=1
        r1=sheet.cell(row=j,column=1).value
        if user == r1:
            r2.add(sheet.cell(row=j,column=2).value)
print(r2)
            
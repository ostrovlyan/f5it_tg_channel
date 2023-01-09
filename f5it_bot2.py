from re import sub
from typing_extensions import get_overloads
from openpyxl import Workbook, load_workbook

import telebot
from telebot import types

wb = Workbook()
wb=load_workbook(filename='price.xlsx')
sheet = wb.active

def read_from_file(path: str) -> list[list[str]]:
    wb_obj = load_workbook(path)
    sheet_obj = wb_obj.active
    goods = []
    for row in sheet_obj.iter_rows(min_row=2, min_col=1, max_row=sheet_obj.max_row, max_col=sheet_obj.max_column):
        good = []
        for cell in row:
            good.append(cell.value)
        goods.append(good)

    return goods



#goods = read_from_file("price.xlsx")

def get_categories(goods: list[list[str]]) -> list[str]:
    categories = set()
    for item in goods:
        categories.add(item[0])

    return list(categories)

def get_subcategories(goods: list[list[str]]) -> list[str]:
    subcategories = set()
    for item in goods:
        subcategories.add(item[1])

    return list(subcategories)


result = read_from_file("price.xlsx")
categories = get_categories(result)
subcategories=get_subcategories(result)
#print(result, end='\n')
#print(categories)
#print(subcategories)

r2=set()
r7=set()
j=0



token='token'
bot=telebot.TeleBot(token)
@bot.message_handler(commands=["start"])
def start(message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(*[types.KeyboardButton(name) for name in ['Меню', 'Прайс-лист']])
    keyboard.add(*[types.KeyboardButton(name) for name in ['Предложение дня', 'Заказать']])
    bot.send_message(message.chat.id, 'выберите вариант', reply_markup=keyboard)
@bot.message_handler(content_types=['text'])
def message(message):
    if message.text == 'Заказать':
        bot.send_message(message.chat.id, 'Тут инфа о компании1')
    elif message.text == 'Меню':
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(*[types.KeyboardButton(name) for name in ['Меню', 'Прайс-лист']])
        keyboard.add(*[types.KeyboardButton(name) for name in ['Предложение дня', 'Заказать']])
        bot.send_message(message.chat.id, 'выберите вариант', reply_markup=keyboard)

    elif message.text == 'Прайс-лист':       
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True,row_width=2)
        for cats in categories:
            keyboard.add(cats)
        bot.send_message(message.chat.id, 'Выберите категорию', reply_markup=keyboard)
    elif message.text in categories:
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
        j=0
        for row in sheet.iter_rows(min_row=1, min_col=2, max_row=sheet.max_row, max_col=2):
            for cell in row:
                j+=1
                r1=sheet.cell(row=j,column=1).value
                if message.text == r1:
                    r2.add(sheet.cell(row=j,column=2).value)
        #print(r2)
        for each in r2:
            keyboard.add(each)

        bot.send_message(message.chat.id, 'Выберите подкатегорию', reply_markup=keyboard)

    elif message.text in subcategories:
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
        j=0
        for row in sheet.iter_rows(min_row=1, min_col=7, max_row=sheet.max_row, max_col=7):
            for cell in row:
                j+=1
                r1=sheet.cell(row=j,column=2).value
                if message.text == r1:
                    r7.add(sheet.cell(row=j,column=7).value)
        #print(r2)
        for each in r7:
            bot.send_message(message.chat.id, each, reply_markup=keyboard)
        keyboard.add(*[types.KeyboardButton(name) for name in ['Меню', 'Заказать']])
        bot.send_message(message.chat.id, 'выберите вариант', reply_markup=keyboard)
        # new untested shit
        #for each in subcategories:






bot.infinity_polling()
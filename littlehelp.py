from openpyxl import load_workbook


def read_from_file(path: str) -> list[list[str]]:
    wb_obj = load_workbook(path)
    sheet_obj = wb_obj.active
    goods = []
    for row in sheet_obj.iter_rows(min_row=1, min_col=1, max_row=sheet_obj.max_row, max_col=sheet_obj.max_column):
        good = []
        for cell in row:
            good.append(cell.value)
        goods.append(good)

    return goods


#goods = read_from_file("price.xlsx")

def get_categories(goods: list[list[str]]) -> list[str]:
    categories = set()
    for item in goods:
        categories.add(item[0])

    return list(categories)


result = read_from_file("price.xlsx")
categories = get_categories(result)

print(categories)